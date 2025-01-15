import openpyxl
import os


file = os.path.join(os.getcwd(),"day14\\data.xlsx") #file path of the excel file

workbook=openpyxl.load_workbook(file)
sheet = workbook["Orders"]

rows = sheet.max_row #no of rows in sheet
columns = sheet.max_column #no of columns in sheet

print(rows,columns)

#read data from the sheet

for r in range(1,rows+1):
    for c in range(1,columns+1):
        print(sheet.cell(r,c).value,end=" ") #get value for the exact row and column
    print("\n")



